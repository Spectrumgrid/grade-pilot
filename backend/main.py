from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pandas as pd
import io
import logging
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg') # Usar backend no interactivo
import os
import uuid
from pathlib import Path
import json
import time

ENV = os.getenv("ENV", "dev")
FRONTEND_URL = os.getenv("FRONTEND_URL", "http://localhost:5173")
ROOT_PATH = os.getenv("ROOT_PATH", "")


SESSIONS_DIR = Path("sessions")
SESSIONS_DIR.mkdir(exist_ok=True)
SESSION_MAX_AGE_SECONDS = 60 * 60 * 24 * 365  # 1 año

# Configuración de logs
if not os.path.exists("logs"):
    os.makedirs("logs")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/corrector_backend.log")
    ]
)
logger = logging.getLogger("corrector")

def cleanup_old_sessions():
    now = time.time()
    removed = 0

    for session_dir in SESSIONS_DIR.iterdir():
        if not session_dir.is_dir():
            continue

        try:
            # Usamos la fecha de creación/modificación del directorio
            last_modified = session_dir.stat().st_mtime
            age = now - last_modified

            if age > SESSION_MAX_AGE_SECONDS:
                for file in session_dir.iterdir():
                    file.unlink()
                session_dir.rmdir()
                removed += 1
        except Exception as e:
            logger.warning(f"No se pudo limpiar la sesión {session_dir.name}: {e}")

    if removed > 0:
        logger.info(f"Limpieza de sesiones antiguas: {removed} sesiones eliminadas")

app = FastAPI(root_path=ROOT_PATH)

cleanup_old_sessions()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "https://grade-pilot-1.onrender.com",
        "http://feqxtools.uvigo.es",
        "https://feqxtools.uvigo.es",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


OPCIONES_TOTALES = {"A", "B", "C", "D", "E"}

def calculate_metrics_logic(preview_list):
    if not preview_list:
        return {}

    total_alumnos = len(preview_list)
    presentados = [x for x in preview_list if x["presentado"]]
    no_presentados = total_alumnos - len(presentados)

    if presentados:
        notas_presentados = [x["nota"] for x in presentados]
        media = round(sum(notas_presentados) / len(notas_presentados), 2)
        nota_max = round(max(notas_presentados), 2)
        nota_min = round(min(notas_presentados), 2)
        aprobados = sum(1 for x in notas_presentados if x >= 5)
        suspensos = len(notas_presentados) - aprobados
        porcentaje_aprobados = round(
            (aprobados / len(notas_presentados)) * 100, 1
        )
    else:
        media = nota_max = nota_min = 0
        aprobados = suspensos = porcentaje_aprobados = 0

    return {
        "alumnos_totales": total_alumnos,
        "presentados": len(presentados),
        "no_presentados": no_presentados,
        "media": media,
        "max": nota_max,
        "min": nota_min,
        "aprobados": aprobados,
        "suspensos": suspensos,
        "porcentaje_aprobados": porcentaje_aprobados,
    }


def parse_respuesta(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return set()
    return set(v.strip().upper() for v in str(valor).split(","))




def validate_excel_logic(df: pd.DataFrame, n_preguntas: int, n_opciones: int):
    """
    Función centralizada para validar un DataFrame de examen.
    Lanza HTTPException(400) si encuentra cualquier error.
    """
    if df.empty:
        logger.error("El archivo Excel está vacío")
        raise HTTPException(status_code=400, detail="El archivo Excel está vacío. Sube un archivo con datos.")

    if len(df.columns) != n_preguntas + 1:
        logger.error(f"Número de columnas incorrecto: {len(df.columns)}")
        raise HTTPException(status_code=400, detail=f"El Excel debe tener exactamente {n_preguntas + 1} columnas (1 para DNI y {n_preguntas} para preguntas). Se detectaron {len(df.columns)}.")

    # VALIDACIÓN: Fila de respuestas correctas
    if df.iloc[0].isna().all():
        logger.error("Fila 2 vacía")
        raise HTTPException(status_code=400, detail="La fila 2 (clave de respuestas) está totalmente vacía. No puedo corregir sin soluciones.")

    LETRAS_POSIBLES = ["A", "B", "C", "D", "E"]
    OPCIONES_ACTUALES = set(LETRAS_POSIBLES[:n_opciones])

    # --- ESCANEO INTELIGENTE DE OPCIONES EN EL EXCEL ---
    all_text_in_questions = []
    for col in range(1, n_preguntas + 1):
        all_text_in_questions.extend(df.iloc[:, col].astype(str).tolist())
    
    letras_detectadas = set()
    for text in all_text_in_questions:
        if pd.isna(text) or text.lower() == 'nan': continue
        for char in text.upper():
            if char in LETRAS_POSIBLES:
                letras_detectadas.add(char)
    
    if letras_detectadas:
        max_letra_detectada = max(letras_detectadas)
        idx_max_detectada = LETRAS_POSIBLES.index(max_letra_detectada) + 1
        
        if idx_max_detectada > n_opciones:
            logger.error(f"Mismatch opciones: detectada {max_letra_detectada} pero configuradas {n_opciones}")
            raise HTTPException(
                status_code=400, 
                detail=f"Has configurado el examen con {n_opciones} opciones, pero he detectado respuestas con la letra '{max_letra_detectada}'. Por favor, selecciona {idx_max_detectada} opciones en el menú desplegable."
            )
        
        if idx_max_detectada < n_opciones:
            logger.error(f"Mismatch opciones: detectada max {max_letra_detectada} pero configuradas {n_opciones}")
            raise HTTPException(
                status_code=400,
                detail=f"Parece que el examen es de {idx_max_detectada} opciones (la letra más alta es '{max_letra_detectada}'), pero has seleccionado {n_opciones}. Esto causaría errores en el cálculo de la penalización. Por favor, corrígelo."
            )

    # Validar clave
    for i in range(n_preguntas):
        col_idx = i + 1
        original_val = df.iloc[0, col_idx]
        clave_set = parse_respuesta(original_val)
        
        if not clave_set:
            logger.error(f"Pregunta P{i+1} sin clave")
            raise HTTPException(status_code=400, detail=f"La pregunta P{i+1} en la CLAVE (fila 2) no tiene ninguna respuesta seleccionada.")

        if not clave_set.issubset(OPCIONES_ACTUALES):
            opciones_invalidas = clave_set - OPCIONES_ACTUALES
            logger.error(f"Opción inválida en clave P{i+1}: {opciones_invalidas}")
            raise HTTPException(
                status_code=400, 
                detail=f"En la P{i+1} de la CLAVE (fila 2) has puesto '{','.join(opciones_invalidas)}', pero has configurado el examen para tener solo {n_opciones} opciones."
            )

    df_alumnos = df.iloc[1:].copy()
    if df_alumnos.empty:
        logger.error("No hay alumnos")
        raise HTTPException(status_code=400, detail="No hay datos de alumnos. Asegúrate de listarlos a partir de la fila 3.")

    if df_alumnos.iloc[:, 0].isna().all():
        logger.error("Columna DNI totalmente vacía")
        raise HTTPException(status_code=400, detail="La columna de DNI está vacía. No puedo identificar a los alumnos.")


@app.post("/corregir")
async def corregir_examen(
    file: UploadFile = File(...),
    n_opciones: int = Form(5),
    n_preguntas: int = Form(10)
):
    session_id = str(uuid.uuid4())
    session_path = SESSIONS_DIR / session_id
    session_path.mkdir()

    logger.info(f"Iniciando corrección de archivo: {file.filename} con {n_opciones} opciones y {n_preguntas} preguntas")
    
    try:
        if not (5 <= n_preguntas <= 20):
            raise HTTPException(status_code=400, detail="El número de preguntas debe estar entre 5 y 20")

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")

        try:
            df_original = pd.read_excel(file.file)
        except Exception as e:
            logger.error(f"Error al leer el Excel: {str(e)}")
            raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel. Asegúrate de que no esté corrupto.")

        # VALIDACIÓN COMPLETA
        validate_excel_logic(df_original, n_preguntas, n_opciones)

        LETRAS_POSIBLES = ["A", "B", "C", "D", "E"]
        OPCIONES_ACTUALES = set(LETRAS_POSIBLES[:n_opciones])

        respuestas_correctas = {}
        for i in range(n_preguntas):
            col_idx = i + 1
            original_val = df_original.iloc[0, col_idx]
            respuestas_correctas[f"P{i+1}"] = parse_respuesta(original_val)

        df_alumnos = df_original.iloc[1:].copy()

    except HTTPException as he:
        raise he
    except Exception as e:
        logger.exception("Error crítico en corrección")
        raise HTTPException(status_code=500, detail=f"Error inesperado al procesar el Excel: {str(e)}")

    notas = []
    preview = []

    for idx, row in df_alumnos.iterrows():
        puntaje_total = 0.0
        presentado = False

        # Sanitizar DNI para evitar errores de JSON (nan)
        dni_alumno = row.iloc[0]
        if pd.isna(dni_alumno):
            dni_alumno = "S/DNI"
        else:
            dni_alumno = str(dni_alumno)

        for i, correctas in enumerate(respuestas_correctas.values()):
            incorrectas = OPCIONES_ACTUALES - correctas
            seleccionadas = parse_respuesta(row.iloc[i + 1])

            if seleccionadas:
                presentado = True
            else:
                continue

            n_correctas = len(correctas)
            n_incorrectas = len(incorrectas)

            valor_correcta = 1 / n_correctas if n_correctas else 0
            valor_incorrecta = 1 / n_incorrectas if n_incorrectas else 0

            for opcion in seleccionadas:
                if opcion in correctas:
                    puntaje_total += valor_correcta
                elif opcion in incorrectas:
                    puntaje_total -= valor_incorrecta

        nota_final = round(max(0, puntaje_total), 2)
        notas.append(nota_final)

        preview.append({
            "dni": dni_alumno,
            "nota": nota_final,
            "presentado": presentado
        })

    metrics = calculate_metrics_logic(preview)
    
    # Calcular estadísticas por pregunta
    question_stats = []
    
    total_alumnos = len(df_alumnos)
    
    for i in range(n_preguntas):
        question_num = i + 1
        correctas = list(respuestas_correctas.values())[i]
        
        # Estadísticas por pregunta (puntuación media)
        total_answered = 0
        total_score = 0.0
        
        # Estadísticas por opción: contar aciertos
        # Acierto = (opción correcta Y marcada) O (opción incorrecta Y NO marcada)
        option_aciertos = {opt: 0 for opt in OPCIONES_ACTUALES}
        
        for _, row in df_alumnos.iterrows():
            seleccionadas = parse_respuesta(row.iloc[i + 1])
            
            # Calcular puntuación de esta pregunta para este alumno
            if seleccionadas:
                total_answered += 1
                
                incorrectas = OPCIONES_ACTUALES - correctas
                n_correctas = len(correctas)
                n_incorrectas = len(incorrectas)
                valor_correcta = 1 / n_correctas if n_correctas else 0
                valor_incorrecta = 1 / n_incorrectas if n_incorrectas else 0
                
                question_score = 0.0
                for opcion in seleccionadas:
                    if opcion in correctas:
                        question_score += valor_correcta
                    elif opcion in incorrectas:
                        question_score -= valor_incorrecta
                
                total_score += max(0, question_score)
            
            # Calcular aciertos por opción para TODOS los alumnos
            for opt in OPCIONES_ACTUALES:
                esta_en_clave = opt in correctas
                esta_marcada = opt in seleccionadas
                acierto = (esta_en_clave and esta_marcada) or (not esta_en_clave and not esta_marcada)
                if acierto:
                    option_aciertos[opt] += 1
        
        avg_score = round(total_score / total_answered, 2) if total_answered > 0 else 0
        
        # Preparar datos de opciones con porcentaje de aciertos
        option_data = {}
        for opt in sorted(OPCIONES_ACTUALES):
            aciertos = option_aciertos[opt]
            porcentaje = round((aciertos / total_alumnos * 100), 1) if total_alumnos > 0 else 0
            option_data[opt] = porcentaje
        
        question_stats.append({
            "question": f"P{question_num}",
            "avg_score": avg_score,
            "total_answered": total_answered,
            **option_data  # A, B, C, D, E como claves directas
        })
    
    
    df_corregido = df_alumnos.copy()
    df_corregido["Nota"] = notas

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_original.to_excel(writer, sheet_name="ORIGINAL", index=False)
        df_corregido.to_excel(writer, sheet_name="CORREGIDO", index=False)
        
        # Nueva hoja de MÉTRICAS
        metrics_data = [
            ["Métrica", "Valor"],
            ["Alumnos totales", metrics["alumnos_totales"]],
            ["Presentados", metrics["presentados"]],
            ["No presentados", metrics["no_presentados"]],
            ["Nota media", metrics["media"]],
            ["Nota máxima", metrics["max"]],
            ["Nota mínima", metrics["min"]],
            ["Aprobados", metrics["aprobados"]],
            ["Suspensos", metrics["suspensos"]],
            ["% Aprobados", f"{metrics['porcentaje_aprobados']}%"],
        ]
        df_metrics = pd.DataFrame(metrics_data[1:], columns=metrics_data[0])
        df_metrics.to_excel(writer, sheet_name="MÉTRICAS", index=False)
        
        # Añadir gráficos de barras
        workbook = writer.book
        worksheet = writer.sheets["MÉTRICAS"]
        from openpyxl.chart import BarChart, Reference
        
        # Preparar datos para los gráficos
        # 1. Puntuación media por pregunta
        question_chart_data = [["Pregunta", "Puntuación media"]]
        for q_data in question_stats:
            question_chart_data.append([q_data["question"], q_data["avg_score"]])
        
        # Escribir datos del gráfico de preguntas
        start_row = len(df_metrics) + 4
        for i, row_data in enumerate(question_chart_data):
            for j, value in enumerate(row_data):
                worksheet.cell(row=start_row + i, column=j + 1, value=value)
        
        # Crear gráfico de puntuación media por pregunta
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Puntuación media por pregunta"
        chart1.y_axis.title = "Puntuación"
        chart1.x_axis.title = "Pregunta"
        
        data1 = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row + len(question_stats))
        cats1 = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(question_stats))
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(cats1)
        
        worksheet.add_chart(chart1, "H2")
        
        # 2. Porcentaje de aciertos por opción
        option_chart_data = [["Pregunta", "A", "B", "C", "D", "E"]]
        for q_data in question_stats:
            option_chart_data.append([
                q_data["question"],
                q_data.get("A", 0),
                q_data.get("B", 0),
                q_data.get("C", 0),
                q_data.get("D", 0),
                q_data.get("E", 0)
            ])
        
        # Escribir datos del gráfico de opciones
        start_row2 = start_row + len(question_chart_data) + 3
        for i, row_data in enumerate(option_chart_data):
            for j, value in enumerate(row_data):
                worksheet.cell(row=start_row2 + i, column=j + 1, value=value)
        
        # Crear gráfico de porcentaje de aciertos por opción
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 11
        chart2.title = "Porcentaje de aciertos por opción (%)"
        chart2.y_axis.title = "% Aciertos"
        chart2.x_axis.title = "Pregunta"
        chart2.grouping = "clustered"
        
        data2 = Reference(worksheet, min_col=2, min_row=start_row2, max_row=start_row2 + len(question_stats), max_col=6)
        cats2 = Reference(worksheet, min_col=1, min_row=start_row2 + 1, max_row=start_row2 + len(question_stats))
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        
        worksheet.add_chart(chart2, "H18")

    excel_bytes = output.getvalue()
    
    # Guardar preview
    with open(session_path / "preview.json", "w", encoding="utf-8") as f:
        json.dump(preview, f, ensure_ascii=False)

    # Guardar métricas completas
    metrics_full = metrics.copy()
    metrics_full["question_data"] = question_stats
    metrics_full["n_opciones"] = n_opciones

    with open(session_path / "metrics.json", "w", encoding="utf-8") as f:
        json.dump(metrics_full, f, ensure_ascii=False)

    # Guardar Excel
    with open(session_path / "examen.xlsx", "wb") as f:
        f.write(excel_bytes)

    
    return {"session_id": session_id}

@app.post("/validate")
async def validar_examen(
    file: UploadFile = File(...),
    n_opciones: int = Form(5),
    n_preguntas: int = Form(10)
):
    logger.info(f"Validación previa del archivo: {file.filename} con {n_opciones} opciones y {n_preguntas} preguntas")

    try:
        if not (5 <= n_preguntas <= 20):
            raise HTTPException(400, "El número de preguntas debe estar entre 5 y 20")

        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(400, "El archivo debe ser un Excel (.xlsx o .xls)")

        try:
            df = pd.read_excel(file.file)
        except Exception:
            raise HTTPException(400, "No se pudo leer el archivo Excel")

        # VALIDACIÓN CENTRALIZADA
        validate_excel_logic(df, n_preguntas, n_opciones)

        return {
            "status": "ok",
            "message": "El archivo es válido y puede corregirse"
        }

    except HTTPException as e:
        raise e
    except Exception as e:
        logger.exception("Error en validación previa")
        raise HTTPException(500, "Error inesperado durante la validación")



@app.get("/download/{session_id}")
def download_excel(session_id: str):
    path = SESSIONS_DIR / session_id / "examen.xlsx"
    if not path.exists():
        raise HTTPException(404, "Excel no encontrado")

    return StreamingResponse(
        open(path, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=examen_corregido.xlsx"}
    )

@app.get("/export-pdf/{session_id}")
async def export_pdf(session_id: str):
    session_path = SESSIONS_DIR / session_id

    preview_path = session_path / "preview.json"
    metrics_path = session_path / "metrics.json"

    if not preview_path.exists() or not metrics_path.exists():
        raise HTTPException(404, "Datos no encontrados para esta sesión")

    preview = json.loads(preview_path.read_text(encoding="utf-8"))
    metrics = json.loads(metrics_path.read_text(encoding="utf-8"))
    question_data = metrics.get("question_data", [])

    if not preview:
        raise HTTPException(400, "No hay datos para exportar")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # ===== TÍTULO =====
    elements.append(
        Paragraph("Calificaciones", styles["Title"])
    )
    elements.append(Spacer(1, 12))
    
    # ===== TABLA DE ALUMNOS =====    
    data = [["DNI", "Nota"]]
    for row in preview:
        data.append([row["dni"], f"{row['nota']:.2f}"])
    
    t = Table(data, hAlign='CENTER')
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(t)
    
    # SALTO DE PÁGINA
    elements.append(PageBreak())
    
    elements.append(
        Paragraph("Métricas", styles["Title"])
    )
    elements.append(Spacer(1, 12))

    # ===== TABLA DE MÉTRICAS =====
    metrics_data = [
        ["Métrica", "Valor"],
        ["Alumnos Totales", str(metrics["alumnos_totales"])],
        ["Presentados", str(metrics["presentados"])],
        ["No Presentados", str(metrics["no_presentados"])],
        ["Nota Media", str(metrics["media"])],
        ["Nota Máxima", str(metrics["max"])],
        ["Nota Mínima", str(metrics["min"])],
        ["Aprobados", str(metrics["aprobados"])],
        ["Suspensos", str(metrics["suspensos"])],
        ["% Aprobados", f"{metrics['porcentaje_aprobados']}%"]
    ]
    
    tm = Table(metrics_data, hAlign='CENTER')
    tm.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(tm)
    elements.append(Spacer(1, 20))

    # ===== GRÁFICOS EN EL PDF =====
    if question_data:
        # 1. Gráfico de Puntuación Media
        plt.figure(figsize=(6, 4))
        questions = [q["question"] for q in question_data]
        avg_scores = [q["avg_score"] for q in question_data]
        
        plt.bar(questions, avg_scores, color='#45B7D1')
        plt.title("Puntuación media por pregunta")
        plt.ylim(0, 1.1)
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        img_buffer1 = io.BytesIO()
        plt.savefig(img_buffer1, format='png', bbox_inches='tight')
        img_buffer1.seek(0)
        plt.close()
        
        elements.append(Image(img_buffer1, width=400, height=250))
        elements.append(Spacer(1, 20))
        
        # 2. Gráfico de Aciertos por Opción (Solo si hay datos frescos)
        plt.figure(figsize=(8, 5))
        x = range(len(questions))
        width = 0.15
        
        # Obtener qué opciones existen realmente (A, B, C, D, E)
        OPCIONES = ["A", "B", "C", "D", "E"]
        colors_list = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8']
        
        for idx, opt in enumerate(OPCIONES):
            # Solo pintar si al menos una pregunta tiene esta opción (evita errores si n_opciones < 5)
            vals = [q.get(opt, 0) for q in question_data]
            if any(v > 0 for v in vals):
                plt.bar([p + (idx * width) for p in x], vals, width, label=opt, color=colors_list[idx])
        
        plt.title("Porcentaje de aciertos por opción (%)")
        plt.xticks([p + 2 * width for p in x], questions)
        plt.ylim(0, 105)
        plt.legend()
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        img_buffer2 = io.BytesIO()
        plt.savefig(img_buffer2, format='png', bbox_inches='tight')
        img_buffer2.seek(0)
        plt.close()
        
        elements.append(Image(img_buffer2, width=450, height=280))

    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=reporte_examen.pdf"
        }
    )

@app.get("/preview/{session_id}")
def get_preview(session_id: str):
    path = SESSIONS_DIR / session_id / "preview.json"
    if not path.exists():
        raise HTTPException(404, "Sesión no encontrada")
    return json.loads(path.read_text(encoding="utf-8"))

@app.get("/metrics/{session_id}")
def get_metrics(session_id: str):
    path = SESSIONS_DIR / session_id / "metrics.json"
    if not path.exists():
        raise HTTPException(404, "Sesión no encontrada")
    return json.loads(path.read_text(encoding="utf-8"))

@app.get("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    import os

    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 8000))
    )

